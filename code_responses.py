#!/usr/bin/env python3
"""Code self-disclosure responses along 28 dimensions using Claude Opus 4.7.

Usage:
    export ANTHROPIC_API_KEY=sk-...
    python code_responses.py                  # code all rows
    python code_responses.py --limit 10       # pilot: first 10 rows per study
    python code_responses.py --workers 10     # set concurrency (default 5)
    python code_responses.py --dry-run        # no API calls; fill with placeholders
    python code_responses.py --study 3        # only Study 3

Resume is automatic: completed rows (keyed by Study/ID) are written to
.coding_checkpoint.jsonl and skipped on re-run. Delete that file to start over.
"""

from __future__ import annotations

import argparse
import json
import os
import random
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import anthropic
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from pydantic import BaseModel, Field

from rubric import RUBRIC

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

MODEL = "claude-opus-4-7"
INPUT_FILE = "Open Responses.xlsx"
OUTPUT_FILE = "Open Responses - Coded.xlsx"
CHECKPOINT_FILE = ".coding_checkpoint.jsonl"

SHEETS = ["STUDY 1", "STUDY 2", "STUDY 3"]

# 28 output columns in the exact order specified by §8 of the instructions.
OUTPUT_COLUMNS = [
    "SCORE 1", "SCORE 2",
    "Sentiment", "Score_S",
    "Content", "Score_C",
    "Topic", "Score_T",
    "Narrative_Style", "Score_NS",
    "Resolution", "Score_R",
    "Focus", "Score_F",
    "Self_Censorship", "IC_Language", "Spontaneous_Revelation",
    "Realtime_Processing", "Emotional_Regulation", "Emotional_Spillover",
    "Other_Self", "Score_OS",
    "Topic_Sensitivity", "Score_TS",
    "Temporal_Orientation", "Score_TO",
    "Target_Closeness", "Score_TC",
]

# Score_* -> allowed string label mappings (for validation).
SENTIMENT_MAP = {
    0: "not applicable", 1: "negative", 2: "positive",
    3: "neutral/mixed", 4: "unresolved",
}
NARRATIVE_MAP = {0: "not applicable", 1: "sequential", 2: "summary", 3: "hybrid"}
RESOLUTION_MAP = {0: "not mentioned", 1: "resolved", 2: "partial", 3: "unresolved"}
FOCUS_MAP = {0: "not applicable", 1: "situation", 2: "self", 3: "balanced"}
OS_MAP = {0: "not applicable", 1: "other-oriented", 2: "self-oriented", 3: "balanced"}
TS_MAP = {0: "not applicable", 1: "low", 2: "medium", 3: "high"}
TO_MAP = {0: "not applicable", 1: "past-resolved", 2: "ongoing", 3: "mixed"}
TC_MAP = {
    0: "not applicable", 1: "stranger", 2: "acquaintance",
    3: "friend", 4: "family", 5: "partner",
}

# ---------------------------------------------------------------------------
# Output schema (Pydantic model used for structured output)
# ---------------------------------------------------------------------------


class CodedResponse(BaseModel):
    """The full 28-column coding for one response, plus a short rationale."""

    rationale: str = Field(
        description=(
            "One to three sentences explaining the key coding decisions, "
            "especially any borderline calls."
        )
    )

    score_1: int = Field(ge=0, le=3, description="Rubric 1 (0-3)")
    score_2: int = Field(ge=0, le=6, description="Rubric 2 (0-6)")

    sentiment: Literal[
        "negative", "positive", "neutral/mixed", "unresolved", "not applicable"
    ]
    score_s: int = Field(ge=0, le=4)

    content: str = Field(description="<=15 word description of what the response is about")
    score_c: int = Field(ge=0, le=6)

    topic: str = Field(description="1-3 word topic domain label")
    score_t: int = Field(ge=0, le=11)

    narrative_style: Literal[
        "sequential", "summary", "hybrid", "not applicable"
    ]
    score_ns: int = Field(ge=0, le=3)

    resolution: Literal[
        "resolved", "partial", "unresolved", "not mentioned", "not applicable"
    ]
    score_r: int = Field(ge=0, le=3)

    focus: Literal["situation", "self", "balanced", "not applicable"]
    score_f: int = Field(ge=0, le=3)

    self_censorship: int = Field(ge=0, le=1)
    ic_language: int = Field(ge=0, le=1)
    spontaneous_revelation: int = Field(ge=0, le=1)
    realtime_processing: int = Field(ge=0, le=1)
    emotional_regulation: int = Field(ge=0, le=1)
    emotional_spillover: int = Field(ge=0, le=1)

    other_self: Literal[
        "other-oriented", "self-oriented", "balanced", "not applicable"
    ]
    score_os: int = Field(ge=0, le=3)

    topic_sensitivity: Literal["low", "medium", "high", "not applicable"]
    score_ts: int = Field(ge=0, le=3)

    temporal_orientation: Literal[
        "past-resolved", "ongoing", "mixed", "not applicable"
    ]
    score_to: int = Field(ge=0, le=3)

    target_closeness: Literal[
        "stranger", "acquaintance", "friend", "family", "partner", "not applicable"
    ]
    score_tc: int = Field(ge=0, le=5)


# ---------------------------------------------------------------------------
# Prompt
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = f"""{RUBRIC}

## Output

Return structured output with all 28 coded fields and a short rationale.
Before emitting, verify:

  - The R1/R2 consistency rule (see Critical rules §4) holds.
  - If Self_Censorship = 1 then IC_Language = 1.
  - For blank / invalid responses, all integer scores are 0 and all string
    fields are "not applicable".
  - Each Score_* integer matches its paired string column per the rubric
    tables.

If a value is genuinely borderline, note it briefly in `rationale`.
"""

# ---------------------------------------------------------------------------
# Checkpoint
# ---------------------------------------------------------------------------

_checkpoint_lock = threading.Lock()


def load_checkpoint(path: Path) -> dict[tuple[str, str], dict]:
    if not path.exists():
        return {}
    out: dict[tuple[str, str], dict] = {}
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            out[(rec["sheet"], str(rec["id"]))] = rec
    return out


def append_checkpoint(path: Path, rec: dict) -> None:
    with _checkpoint_lock:
        with path.open("a") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ---------------------------------------------------------------------------
# Validation / post-processing
# ---------------------------------------------------------------------------


@dataclass
class ValidationIssue:
    sheet: str
    row_id: str
    field: str
    before: object
    after: object
    message: str


def validate_and_normalise(
    coded: CodedResponse, *, sheet: str, row_id: str
) -> tuple[CodedResponse, list[ValidationIssue]]:
    """Fix obvious consistency violations and collect notes for review."""

    issues: list[ValidationIssue] = []
    data = coded.model_dump()

    # R1 / R2 consistency
    allowed_r2 = {0: {0}, 1: {1, 2}, 2: {3, 4}, 3: {5, 6}}
    if data["score_2"] not in allowed_r2[data["score_1"]]:
        before = data["score_2"]
        # snap R2 into the allowed band (closest value)
        data["score_2"] = min(allowed_r2[data["score_1"]], key=lambda x: abs(x - before))
        issues.append(ValidationIssue(
            sheet, row_id, "score_2", before, data["score_2"],
            f"R1={data['score_1']} requires R2 in {sorted(allowed_r2[data['score_1']])}",
        ))

    # IC_Language must be >= Self_Censorship
    if data["self_censorship"] == 1 and data["ic_language"] == 0:
        issues.append(ValidationIssue(
            sheet, row_id, "ic_language", 0, 1,
            "IC_Language must be 1 when Self_Censorship is 1",
        ))
        data["ic_language"] = 1

    # Paired string/code mismatches: snap string to match code.
    def snap(field: str, code_field: str, mapping: dict[int, str], *,
             blank_override: str | None = None) -> None:
        code = data[code_field]
        expected = mapping.get(code)
        if blank_override is not None and code == 0:
            expected = blank_override
        if expected is not None and data[field] != expected:
            issues.append(ValidationIssue(
                sheet, row_id, field, data[field], expected,
                f"string did not match {code_field}={code}",
            ))
            data[field] = expected

    snap("sentiment", "score_s", SENTIMENT_MAP)
    snap("narrative_style", "score_ns", NARRATIVE_MAP)
    # Resolution: blank rows use "not applicable", otherwise use the code label.
    # We can't tell here if the row is "blank" (no other signal), so trust the
    # model unless score_r is out of range or clearly contradicts.
    if data["resolution"] not in (
        "resolved", "partial", "unresolved", "not mentioned", "not applicable"
    ):
        issues.append(ValidationIssue(
            sheet, row_id, "resolution", data["resolution"], "not mentioned",
            "invalid resolution label",
        ))
        data["resolution"] = "not mentioned"
    snap("focus", "score_f", FOCUS_MAP)
    snap("other_self", "score_os", OS_MAP)
    snap("topic_sensitivity", "score_ts", TS_MAP)
    snap("temporal_orientation", "score_to", TO_MAP)
    snap("target_closeness", "score_tc", TC_MAP)

    # Blank-row invariant: if SCORE 1 = 0, zero-out binaries and set blanket strings.
    if data["score_1"] == 0:
        for k in [
            "score_2", "score_s", "score_c", "score_t",
            "score_ns", "score_r", "score_f",
            "score_os", "score_ts", "score_to", "score_tc",
            "self_censorship", "ic_language", "spontaneous_revelation",
            "realtime_processing", "emotional_regulation", "emotional_spillover",
        ]:
            if data[k] != 0:
                issues.append(ValidationIssue(
                    sheet, row_id, k, data[k], 0, "SCORE 1 = 0 forces all scores to 0",
                ))
                data[k] = 0
        for k in [
            "sentiment", "narrative_style", "resolution", "focus",
            "other_self", "topic_sensitivity", "temporal_orientation",
            "target_closeness",
        ]:
            if data[k] != "not applicable":
                issues.append(ValidationIssue(
                    sheet, row_id, k, data[k], "not applicable",
                    "SCORE 1 = 0 forces string columns to 'not applicable'",
                ))
                data[k] = "not applicable"

    # Content / Topic truncation
    if len(data["content"].split()) > 15:
        truncated = " ".join(data["content"].split()[:15])
        issues.append(ValidationIssue(
            sheet, row_id, "content", data["content"], truncated,
            "content exceeded 15 words",
        ))
        data["content"] = truncated
    if len(data["topic"].split()) > 3:
        truncated = " ".join(data["topic"].split()[:3])
        issues.append(ValidationIssue(
            sheet, row_id, "topic", data["topic"], truncated,
            "topic exceeded 3 words",
        ))
        data["topic"] = truncated

    return CodedResponse.model_validate(data), issues


# ---------------------------------------------------------------------------
# API call
# ---------------------------------------------------------------------------


def code_one(
    client: anthropic.Anthropic,
    response_text: str,
    *,
    max_retries: int = 4,
) -> CodedResponse:
    """Call the API once and return a validated CodedResponse."""
    last_err: Exception | None = None
    for attempt in range(max_retries):
        try:
            msg = client.messages.parse(
                model=MODEL,
                max_tokens=2048,
                system=[
                    {
                        "type": "text",
                        "text": SYSTEM_PROMPT,
                        "cache_control": {"type": "ephemeral"},
                    }
                ],
                messages=[
                    {
                        "role": "user",
                        "content": (
                            "Code the following response. Apply the rubric "
                            "strictly. The response may contain verbal filler "
                            "if it was spoken - ignore filler when judging "
                            "content depth.\n\n"
                            f"<response>\n{response_text}\n</response>"
                        ),
                    }
                ],
                output_format=CodedResponse,
            )
            return msg.parsed_output
        except (anthropic.RateLimitError, anthropic.APIConnectionError,
                anthropic.InternalServerError) as exc:
            last_err = exc
            sleep_s = 2 ** attempt + random.uniform(0, 1)
            time.sleep(sleep_s)
        except anthropic.APIStatusError as exc:
            if 500 <= exc.status_code < 600:
                last_err = exc
                time.sleep(2 ** attempt + random.uniform(0, 1))
                continue
            raise
    assert last_err is not None
    raise last_err


def placeholder_coding() -> CodedResponse:
    """Build a deterministic placeholder for --dry-run."""
    return CodedResponse(
        rationale="DRY RUN placeholder; not a real coding.",
        score_1=0, score_2=0,
        sentiment="not applicable", score_s=0,
        content="DRY RUN", score_c=0,
        topic="n/a", score_t=0,
        narrative_style="not applicable", score_ns=0,
        resolution="not applicable", score_r=0,
        focus="not applicable", score_f=0,
        self_censorship=0, ic_language=0, spontaneous_revelation=0,
        realtime_processing=0, emotional_regulation=0, emotional_spillover=0,
        other_self="not applicable", score_os=0,
        topic_sensitivity="not applicable", score_ts=0,
        temporal_orientation="not applicable", score_to=0,
        target_closeness="not applicable", score_tc=0,
    )


# ---------------------------------------------------------------------------
# Workbook I/O
# ---------------------------------------------------------------------------


def load_responses(path: Path) -> list[tuple[str, str, str]]:
    """Return a list of (sheet, id, response_text)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[tuple[str, str, str]] = []
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"warn: sheet {sheet_name!r} not found, skipping", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            id_col = header.index("ID")
            resp_col = header.index("RESPONSE")
        except ValueError:
            print(f"warn: {sheet_name} missing ID/RESPONSE columns", file=sys.stderr)
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_id = row[id_col]
            resp = row[resp_col]
            if row_id is None:
                continue
            rows.append((sheet_name, str(row_id), str(resp) if resp is not None else ""))
    return rows


def write_output(
    input_path: Path,
    output_path: Path,
    records: dict[tuple[str, str], dict],
) -> None:
    """Add 28 coded columns to each study sheet + append a 'Rationale' sheet."""
    wb = load_workbook(input_path)
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            id_col_idx = header.index("ID")
        except ValueError:
            continue

        # Extend header to exactly: [ID, RESPONSE, <28 OUTPUT_COLUMNS>].
        # The input already has "SCORE 1" and "SCORE 2" headers; overwrite
        # everything from column 3 onward.
        existing_cols = ws.max_column
        start_col = 3  # columns 1=ID, 2=RESPONSE
        for idx, col_name in enumerate(OUTPUT_COLUMNS):
            ws.cell(row=1, column=start_col + idx, value=col_name)
        # Clear any stale columns past the new last column.
        for extra in range(start_col + len(OUTPUT_COLUMNS), existing_cols + 1):
            ws.cell(row=1, column=extra, value=None)

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            row_id = row[id_col_idx]
            if row_id is None:
                continue
            rec = records.get((sheet_name, str(row_id)))
            if rec is None:
                continue
            values = _record_to_column_values(rec)
            for i, v in enumerate(values):
                ws.cell(row=row_idx, column=start_col + i, value=v)

    # Rationale sheet.
    if "Rationale" in wb.sheetnames:
        del wb["Rationale"]
    ws_r = wb.create_sheet("Rationale")
    ws_r.append(["Study", "ID", "Rationale", "Validation_Issues"])
    for (sheet_name, row_id), rec in sorted(records.items()):
        issues = rec.get("validation_issues") or []
        issues_text = "; ".join(
            f"{i['field']}: {i['before']!r}->{i['after']!r} ({i['message']})"
            for i in issues
        )
        ws_r.append([sheet_name, row_id, rec.get("rationale", ""), issues_text])

    wb.save(output_path)


def _record_to_column_values(rec: dict) -> list:
    """Map a checkpoint record to 28 values in OUTPUT_COLUMNS order."""
    c = rec["coded"]
    return [
        c["score_1"], c["score_2"],
        c["sentiment"], c["score_s"],
        c["content"], c["score_c"],
        c["topic"], c["score_t"],
        c["narrative_style"], c["score_ns"],
        c["resolution"], c["score_r"],
        c["focus"], c["score_f"],
        c["self_censorship"], c["ic_language"], c["spontaneous_revelation"],
        c["realtime_processing"], c["emotional_regulation"], c["emotional_spillover"],
        c["other_self"], c["score_os"],
        c["topic_sensitivity"], c["score_ts"],
        c["temporal_orientation"], c["score_to"],
        c["target_closeness"], c["score_tc"],
    ]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, default=None,
                        help="Code only the first N rows per study (pilot mode).")
    parser.add_argument("--workers", type=int, default=5,
                        help="Concurrent API calls (default 5).")
    parser.add_argument("--study", type=int, choices=[1, 2, 3], default=None,
                        help="Only process this study.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Skip API calls; fill with placeholder values to test I/O.")
    parser.add_argument("--input", default=INPUT_FILE)
    parser.add_argument("--output", default=OUTPUT_FILE)
    parser.add_argument("--checkpoint", default=CHECKPOINT_FILE)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    checkpoint_path = Path(args.checkpoint)

    if not input_path.exists():
        print(f"error: {input_path} not found", file=sys.stderr)
        return 2

    all_rows = load_responses(input_path)
    if args.study is not None:
        target_sheet = f"STUDY {args.study}"
        all_rows = [r for r in all_rows if r[0] == target_sheet]

    if args.limit is not None:
        by_study: dict[str, list] = {}
        for r in all_rows:
            by_study.setdefault(r[0], []).append(r)
        all_rows = []
        for sheet in SHEETS:
            all_rows.extend(by_study.get(sheet, [])[: args.limit])

    # Resume
    checkpoint = load_checkpoint(checkpoint_path)
    todo = [r for r in all_rows if (r[0], r[1]) not in checkpoint]
    print(
        f"{len(all_rows)} rows selected, {len(checkpoint)} already in checkpoint, "
        f"{len(todo)} to code"
    )

    client: anthropic.Anthropic | None = None
    if not args.dry_run:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            print("error: ANTHROPIC_API_KEY not set", file=sys.stderr)
            return 2
        client = anthropic.Anthropic()

    done_count = 0
    fail_count = 0
    issues_total: list[ValidationIssue] = []
    start = time.monotonic()

    def process(sheet: str, row_id: str, text: str) -> dict | None:
        try:
            if args.dry_run:
                coded = placeholder_coding()
            else:
                assert client is not None
                if not text.strip():
                    # Cheap blank-rule shortcut.
                    coded = placeholder_coding()
                    coded.content = "Blank"
                else:
                    coded = code_one(client, text)
            coded, issues = validate_and_normalise(coded, sheet=sheet, row_id=row_id)
            rec = {
                "sheet": sheet,
                "id": row_id,
                "coded": coded.model_dump(),
                "rationale": coded.rationale,
                "validation_issues": [
                    {
                        "field": i.field,
                        "before": repr(i.before),
                        "after": repr(i.after),
                        "message": i.message,
                    }
                    for i in issues
                ],
            }
            append_checkpoint(checkpoint_path, rec)
            return rec
        except Exception as exc:  # noqa: BLE001
            print(f"fail {sheet}/{row_id}: {exc}", file=sys.stderr)
            return None

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(process, s, rid, txt): (s, rid) for (s, rid, txt) in todo
        }
        for fut in as_completed(futures):
            rec = fut.result()
            if rec is None:
                fail_count += 1
            else:
                done_count += 1
                checkpoint[(rec["sheet"], str(rec["id"]))] = rec
                issues_total.extend(rec["validation_issues"] or [])
            if (done_count + fail_count) % 25 == 0:
                elapsed = time.monotonic() - start
                rate = (done_count + fail_count) / elapsed if elapsed else 0
                print(
                    f"  progress: {done_count} ok, {fail_count} fail, "
                    f"{rate:.1f} rows/s"
                )

    print(f"done: {done_count} ok, {fail_count} fail")
    if issues_total:
        print(f"  {len(issues_total)} validation fixes applied")

    print(f"writing {output_path}")
    write_output(input_path, output_path, checkpoint)
    print("finished")
    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
